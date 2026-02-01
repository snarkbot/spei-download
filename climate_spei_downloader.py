#!/usr/bin/env python3
"""
Climate SPEI Data Downloader

Downloads historical and projected SPEI (Standardized Precipitation Evapotranspiration Index)
climate data for any US county, outputting to an Excel file with three tabs:
1. West-Wide Drought Tracker (1895-present)
2. GridMET (1980-present)
3. Climate Projections (Future scenarios)
"""

import os
import sys
import time
import re
import requests
import pandas as pd
import xarray as xr
import geopandas as gpd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Selenium imports
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from webdriver_manager.firefox import GeckoDriverManager


class CountyLookup:
    """Handles county name to FIPS code and geometry lookup."""

    CENSUS_COUNTIES_URL = "https://www2.census.gov/geo/tiger/GENZ2022/shp/cb_2022_us_county_500k.zip"

    def __init__(self):
        self.counties_gdf = None
        self._load_counties()

    def _load_counties(self):
        """Load US county boundaries from Census TIGER/Line files."""
        cache_dir = Path.home() / ".cache" / "climate_spei"
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_file = cache_dir / "us_counties.gpkg"

        if cache_file.exists():
            print("Loading cached county data...")
            self.counties_gdf = gpd.read_file(cache_file)
        else:
            print("Downloading county boundaries from US Census...")
            self.counties_gdf = gpd.read_file(self.CENSUS_COUNTIES_URL)
            self.counties_gdf.to_file(cache_file, driver="GPKG")
            print("County data cached for future use.")

    def lookup(self, county_name: str, state_name: str) -> dict:
        """
        Look up county by name and state.

        Returns dict with:
            - fips: County FIPS code
            - name: County name
            - state: State name
            - state_fips: State FIPS code
            - centroid: (lon, lat) tuple
            - geometry: County polygon
        """
        # State name to FIPS mapping
        state_fips_map = {
            'Alabama': '01', 'Alaska': '02', 'Arizona': '04', 'Arkansas': '05',
            'California': '06', 'Colorado': '08', 'Connecticut': '09', 'Delaware': '10',
            'District of Columbia': '11', 'Florida': '12', 'Georgia': '13', 'Hawaii': '15',
            'Idaho': '16', 'Illinois': '17', 'Indiana': '18', 'Iowa': '19',
            'Kansas': '20', 'Kentucky': '21', 'Louisiana': '22', 'Maine': '23',
            'Maryland': '24', 'Massachusetts': '25', 'Michigan': '26', 'Minnesota': '27',
            'Mississippi': '28', 'Missouri': '29', 'Montana': '30', 'Nebraska': '31',
            'Nevada': '32', 'New Hampshire': '33', 'New Jersey': '34', 'New Mexico': '35',
            'New York': '36', 'North Carolina': '37', 'North Dakota': '38', 'Ohio': '39',
            'Oklahoma': '40', 'Oregon': '41', 'Pennsylvania': '42', 'Rhode Island': '44',
            'South Carolina': '45', 'South Dakota': '46', 'Tennessee': '47', 'Texas': '48',
            'Utah': '49', 'Vermont': '50', 'Virginia': '51', 'Washington': '53',
            'West Virginia': '54', 'Wisconsin': '55', 'Wyoming': '56'
        }

        # Normalize state name
        state_normalized = state_name.strip().title()
        if state_normalized not in state_fips_map:
            raise ValueError(f"Unknown state: {state_name}")

        state_fips = state_fips_map[state_normalized]

        # Filter by state
        state_counties = self.counties_gdf[self.counties_gdf['STATEFP'] == state_fips]

        if len(state_counties) == 0:
            raise ValueError(f"No counties found for state: {state_name}")

        # Normalize county name for matching
        county_normalized = county_name.strip().lower()
        county_normalized = re.sub(r'\s+county$', '', county_normalized)

        # Find matching county
        match = None
        for idx, row in state_counties.iterrows():
            row_name = row['NAME'].lower()
            if row_name == county_normalized:
                match = row
                break

        if match is None:
            # Try partial match
            for idx, row in state_counties.iterrows():
                row_name = row['NAME'].lower()
                if county_normalized in row_name or row_name in county_normalized:
                    match = row
                    break

        if match is None:
            available = sorted(state_counties['NAME'].tolist())
            raise ValueError(
                f"County '{county_name}' not found in {state_name}. "
                f"Available counties: {', '.join(available[:10])}..."
            )

        # Get centroid
        centroid = match.geometry.centroid

        return {
            'fips': match['STATEFP'] + match['COUNTYFP'],
            'county_fips': match['COUNTYFP'],
            'name': match['NAME'],
            'state': state_normalized,
            'state_fips': state_fips,
            'centroid': (centroid.x, centroid.y),
            'geometry': match.geometry
        }


class WebDriverManager:
    """Manages Selenium WebDriver instances."""

    def __init__(self, headless: bool = True):
        self.headless = headless
        self.driver = None

    def get_driver(self):
        """Get or create a Firefox WebDriver instance."""
        if self.driver is None:
            options = FirefoxOptions()
            if self.headless:
                options.add_argument("--headless")

            print("Initializing Firefox WebDriver...")
            service = FirefoxService(GeckoDriverManager().install())
            self.driver = webdriver.Firefox(service=service, options=options)
            self.driver.set_page_load_timeout(60)

        return self.driver

    def close(self):
        """Close the WebDriver."""
        if self.driver:
            self.driver.quit()
            self.driver = None


class WestWideDroughtTracker:
    """Downloads SPEI data from West-Wide Drought Tracker (1895-present)."""

    # Direct NetCDF download URL for 12-month SPEI, December values
    NETCDF_URL = "https://wrcc.dri.edu/wwdt/data/PRISM/spei12/spei12_12_PRISM.nc"

    def __init__(self, driver_manager: WebDriverManager = None):
        # WebDriver not needed for direct NetCDF download
        self.driver_manager = driver_manager
        self.cache_dir = Path.home() / ".cache" / "climate_spei"
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def download(self, county_info: dict) -> tuple:
        """
        Download historical SPEI data averaged over county polygon.

        Returns tuple of (DataFrame with columns Year/SPEI, error_message or None)
        """
        print("\n--- Downloading from West-Wide Drought Tracker ---")
        geometry = county_info['geometry']

        try:
            # Check for cached NetCDF file
            cache_file = self.cache_dir / "spei12_12_PRISM.nc"

            if not cache_file.exists():
                print("Downloading WWDT SPEI NetCDF file (this may take a few minutes)...")
                import subprocess
                result = subprocess.run(
                    ["curl", "-L", "-o", str(cache_file), self.NETCDF_URL],
                    capture_output=True, text=True, timeout=600
                )
                if result.returncode != 0:
                    raise Exception(f"Download failed: {result.stderr}")
                print("Download complete.")
            else:
                print("Using cached WWDT data file.")

            # Open dataset
            print(f"Extracting data for {county_info['name']} County polygon...")
            ds = xr.open_dataset(cache_file)

            # Find grid cells within county polygon
            grid_cells = self._get_cells_in_polygon(ds, geometry)
            print(f"Found {len(grid_cells)} grid cells within county boundary")

            if not grid_cells:
                raise ValueError("No grid cells found within county polygon")

            # Extract and average data for all cells
            print("Averaging SPEI values across county...")
            df = self._extract_polygon_average(ds, grid_cells)

            ds.close()

            print(f"Retrieved {len(df)} years of data from WWDT (1895-{df['Year'].max()})")
            return df, None

        except Exception as e:
            error_msg = str(e)
            print(f"Error accessing WWDT: {error_msg}")
            print("Generating placeholder data structure...")
            return self._generate_placeholder_data(), error_msg

    def _get_cells_in_polygon(self, ds, geometry) -> list:
        """Find all grid cells whose centers fall within the county polygon."""
        from shapely.geometry import Point

        lats = ds['latitude'].values
        lons = ds['longitude'].values

        # Get bounding box to limit search
        minx, miny, maxx, maxy = geometry.bounds

        # Find cells within polygon
        cells = []
        for lat in lats:
            if lat < miny or lat > maxy:
                continue
            for lon in lons:
                if lon < minx or lon > maxx:
                    continue
                if geometry.contains(Point(lon, lat)):
                    cells.append((lat, lon))

        return cells

    def _extract_polygon_average(self, ds, grid_cells: list) -> pd.DataFrame:
        """Extract SPEI data and average across all grid cells in polygon."""
        import numpy as np

        all_data = []

        for lat, lon in grid_cells:
            data = ds['data'].sel(latitude=lat, longitude=lon, method='nearest')
            df = data.to_dataframe().reset_index()
            df['cell'] = f"{lat},{lon}"
            all_data.append(df)

        # Combine all cells
        combined = pd.concat(all_data, ignore_index=True)

        # Extract year
        combined['Year'] = pd.to_datetime(combined['day']).dt.year

        # Average across all cells for each year
        annual_data = combined.groupby('Year')['data'].mean().reset_index()
        annual_data.columns = ['Year', 'SPEI']

        # Sort by year
        annual_data = annual_data.sort_values('Year').reset_index(drop=True)

        return annual_data

    def _generate_placeholder_data(self) -> pd.DataFrame:
        """Generate placeholder structure for manual data entry."""
        current_year = datetime.now().year
        years = list(range(1895, current_year + 1))
        return pd.DataFrame({
            'Year': years,
            'SPEI': [None] * len(years)
        })


class GridMETDownloader:
    """Downloads SPEI data from GridMET (1980-present) using county polygon averaging."""

    OPENDAP_URL = "http://thredds.northwestknowledge.net/thredds/dodsC/MET/spei/spei1y.nc"

    def download(self, county_info: dict) -> tuple:
        """
        Download GridMET SPEI data averaged over county polygon.

        Returns tuple of (DataFrame with columns Year/SPEI, error_message or None)
        """
        print("\n--- Downloading from GridMET ---")
        geometry = county_info['geometry']
        print(f"Extracting GridMET data for {county_info['name']} County polygon...")

        try:
            print("Opening GridMET OPeNDAP dataset...")
            ds = xr.open_dataset(self.OPENDAP_URL)

            # Find grid cells within county polygon
            grid_cells = self._get_cells_in_polygon(ds, geometry)
            print(f"Found {len(grid_cells)} grid cells within county boundary")

            if not grid_cells:
                raise ValueError("No grid cells found within county polygon")

            # Extract and average data for all cells
            print("Averaging SPEI values across county...")
            annual_data = self._extract_polygon_average(ds, grid_cells, 'spei')

            ds.close()
            print(f"Retrieved {len(annual_data)} years of data from GridMET")
            return annual_data, None

        except Exception as e:
            print(f"Error accessing GridMET: {e}")
            print("Attempting alternative GridMET access method...")
            return self._download_alternative(county_info)

    def _get_cells_in_polygon(self, ds, geometry) -> list:
        """Find all grid cells whose centers fall within the county polygon."""
        from shapely.geometry import Point

        lats = ds['lat'].values
        lons = ds['lon'].values

        # Get bounding box to limit search
        minx, miny, maxx, maxy = geometry.bounds

        # Find cells within polygon
        cells = []
        for lat in lats:
            if lat < miny or lat > maxy:
                continue
            for lon in lons:
                if lon < minx or lon > maxx:
                    continue
                if geometry.contains(Point(lon, lat)):
                    cells.append((lat, lon))

        return cells

    def _extract_polygon_average(self, ds, grid_cells: list, spei_var: str) -> pd.DataFrame:
        """Extract SPEI data and average across all grid cells in polygon."""
        import numpy as np

        all_data = []

        for lat, lon in grid_cells:
            data = ds[spei_var].sel(lat=lat, lon=lon, method='nearest')
            df = data.to_dataframe().reset_index()
            df['cell'] = f"{lat},{lon}"
            all_data.append(df)

        # Combine all cells
        combined = pd.concat(all_data, ignore_index=True)

        # Find time column
        time_col = 'day' if 'day' in combined.columns else 'time'

        # Extract year and month
        combined['year'] = pd.to_datetime(combined[time_col]).dt.year
        combined['month'] = pd.to_datetime(combined[time_col]).dt.month

        # Filter to December
        december_data = combined[combined['month'] == 12]

        # Average across all cells and all December pentads for each year
        annual_data = december_data.groupby('year')[spei_var].mean().reset_index()
        annual_data.columns = ['Year', 'SPEI']

        # Filter to valid years
        annual_data = annual_data[annual_data['Year'] >= 1980]

        return annual_data

    def _download_alternative(self, county_info: dict) -> tuple:
        """Alternative method - fall back to centroid if polygon fails."""
        try:
            lon, lat = county_info['centroid']
            print(f"Falling back to centroid extraction: {lat:.4f}°N, {lon:.4f}°W")

            ds = xr.open_dataset(self.OPENDAP_URL)
            data = ds['spei'].sel(lat=lat, lon=lon, method='nearest')
            df = data.to_dataframe().reset_index()

            df['year'] = pd.to_datetime(df['day']).dt.year
            df['month'] = pd.to_datetime(df['day']).dt.month

            december_data = df[df['month'] == 12]
            annual_data = december_data.groupby('year')['spei'].mean().reset_index()
            annual_data.columns = ['Year', 'SPEI']
            annual_data = annual_data[annual_data['Year'] >= 1980]

            ds.close()
            print(f"Retrieved {len(annual_data)} years from centroid fallback")
            return annual_data, None

        except Exception as e:
            error_msg = str(e)
            print(f"Alternative method failed: {error_msg}")
            return self._generate_placeholder_data(), error_msg

    def _generate_placeholder_data(self) -> pd.DataFrame:
        """Generate placeholder structure for manual data entry."""
        current_year = datetime.now().year
        years = list(range(1980, current_year + 1))
        return pd.DataFrame({
            'Year': years,
            'SPEI': [None] * len(years)
        })


class ClimateProjectionsDownloader:
    """Downloads future SPEI projections from MACA via THREDDS OPeNDAP."""

    # THREDDS server with pre-computed SPEI projections
    THREDDS_BASE = "https://tds-proxy.nkn.uidaho.edu/thredds/dodsC"
    SPEI_MONTHS = 12
    RCP = "rcp45"

    TARGET_MODELS = [
        "CNRM-CM5",
        "CanESM2",
        "HadGEM2-ES365",
        "IPSL-CM5A-MR"
    ]

    def __init__(self, driver_manager: WebDriverManager = None):
        self.driver_manager = driver_manager

    def _get_opendap_url(self, model: str) -> str:
        """Build OPeNDAP URL for a specific model."""
        fname = (f"macav2metdata_{self.SPEI_MONTHS}-month_SPEI_{model}_r1i1p1_"
                 f"{self.RCP}_1950_2099_CONUS_monthly.nc")
        return f"{self.THREDDS_BASE}/MACAV2-SPEI/CIRES-NCCASC-DROUGHT-{self.SPEI_MONTHS}-MONTH/{fname}"

    def download(self, county_info: dict) -> tuple:
        """
        Download future SPEI projections for a county polygon.

        Returns tuple of (DataFrame, list of failed models)
        """
        print("\n--- Downloading Climate Projections (MACA RCP 4.5) ---")
        geometry = county_info['geometry']

        print(f"Extracting SPEI projections for {county_info['name']} County...")
        print(f"Models: {', '.join(self.TARGET_MODELS)}")

        results = {'Year': list(range(2020, 2100))}
        failed_models = []

        for model in self.TARGET_MODELS:
            try:
                print(f"  {model}...", end=" ", flush=True)
                model_data = self._download_model(model, geometry)
                results[model] = model_data
                print(f"OK ({len([v for v in model_data if v is not None])} years)")
            except Exception as e:
                print(f"Failed: {e}")
                results[model] = [None] * len(results['Year'])
                failed_models.append(model)

        df = pd.DataFrame(results)
        success_count = len(self.TARGET_MODELS) - len(failed_models)
        print(f"Retrieved projections for {success_count}/{len(self.TARGET_MODELS)} models (2020-2099)")
        return df, failed_models

    def _download_model(self, model: str, geometry) -> list:
        """Download SPEI data for a single model at county centroid."""
        import numpy as np

        url = self._get_opendap_url(model)
        # Use pydap engine to avoid netCDF4 SSL issues on macOS
        ds = xr.open_dataset(url, engine='pydap')

        try:
            # Use centroid for faster remote access
            centroid = geometry.centroid
            lat_target = centroid.y
            lon_target = centroid.x

            # Select nearest grid point and time range in one operation
            spei = ds['SPEI'].sel(
                lat=lat_target,
                lon=lon_target,
                method='nearest'
            ).sel(time=slice('2020-01-01', '2099-12-31'))

            # Filter to December
            dec_mask = spei['time'].dt.month == 12
            dec_data = spei.where(dec_mask, drop=True)

            return list(dec_data.values)

        finally:
            ds.close()

    def _generate_placeholder_data(self) -> pd.DataFrame:
        """Generate placeholder structure for climate projections."""
        years = list(range(2020, 2100))
        data = {'Year': years}
        for model in self.TARGET_MODELS:
            data[model] = [None] * len(years)
        return pd.DataFrame(data)


class ExcelExporter:
    """Exports data to Excel with multiple tabs."""

    def export(self, wwdt_data: pd.DataFrame, gridmet_data: pd.DataFrame,
               projections_data: pd.DataFrame, output_path: str,
               location_name: str = "Unknown County"):
        """
        Export all data to Excel file with five tabs:
        - Raw data tabs: West-Wide Drought Tracker, GridMET, Climate Projections
        - Formatted tabs: Drought indicators -- Historic, Drought indicators -- Projected
        """
        print(f"\n--- Exporting to Excel ---")

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        # Create workbook
        wb = Workbook()

        # Tab 1: West-Wide Drought Tracker
        ws1 = wb.active
        ws1.title = "West-Wide Drought Tracker"
        self._write_dataframe(ws1, wwdt_data)

        # Tab 2: GridMET
        ws2 = wb.create_sheet("GridMET")
        self._write_dataframe(ws2, gridmet_data)

        # Tab 3: Climate Projections
        ws3 = wb.create_sheet("Climate Projections")
        self._write_dataframe(ws3, projections_data)

        # Tab 4: Drought indicators -- Historic (formatted, starting from 1970)
        ws4 = wb.create_sheet("Drought indicators -- Historic")
        self._write_historic_tab(ws4, wwdt_data, gridmet_data, location_name)

        # Tab 5: Drought indicators -- Projected (formatted)
        ws5 = wb.create_sheet("Drought indicators -- Projected")
        self._write_projected_tab(ws5, projections_data, location_name)

        # Save workbook
        wb.save(output_path)
        print(f"Output saved to: {output_path}")

    def _write_historic_tab(self, ws, wwdt_data: pd.DataFrame,
                            gridmet_data: pd.DataFrame, location_name: str):
        """Write the formatted Historic tab with metadata and data starting at row 16."""
        download_date = datetime.now().strftime('%m/%d/%Y')

        # Metadata rows
        ws['B4'] = 'Data Source: West-Wide Drought Tracker (https://wrcc.dri.edu/wwdt/) and GridMET (http://www.climatologylab.org/gridmet.html)'
        ws['B5'] = f'Download Date: {download_date}'
        ws['B7'] = f'Location (County): {location_name}'
        ws['B9'] = 'Y Variable: 1-Year Standardized Precipitation - Evapotranspiration Index (SPEI-12)'
        ws['B10'] = 'Y Product: PRISM (WWDT 1970-1979) / gridMET (1980+)'

        # Combine WWDT (1970-1979) and GridMET (1980+)
        wwdt_1970_1979 = wwdt_data[(wwdt_data['Year'] >= 1970) & (wwdt_data['Year'] < 1980)].copy()
        wwdt_1970_1979['Source'] = 'West-Wide Drought Tracker'

        gridmet_1980_plus = gridmet_data[gridmet_data['Year'] >= 1980].copy()
        gridmet_1980_plus['Source'] = 'GridMET'

        combined = pd.concat([wwdt_1970_1979, gridmet_1980_plus], ignore_index=True)
        combined = combined.sort_values('Year').reset_index(drop=True)
        combined.columns = ['Year', 'SPEI-12 December', 'Source']
        combined['SPEI-12 December'] = combined['SPEI-12 December'].round(2)

        # Row 15: Headers
        ws['B15'] = 'Year'
        ws['C15'] = 'SPEI-12 December'
        ws['D15'] = 'Source'

        # Row 16+: Data
        for idx, row in combined.iterrows():
            ws[f'B{16 + idx}'] = row['Year']
            ws[f'C{16 + idx}'] = row['SPEI-12 December']
            ws[f'D{16 + idx}'] = row['Source']

    def _write_projected_tab(self, ws, projections_data: pd.DataFrame, location_name: str):
        """Write the formatted Projected tab with metadata and data starting at row 16."""
        download_date = datetime.now().strftime('%m/%d/%Y')

        # Metadata rows
        ws['B4'] = 'Data Source: MACA Climate Projections (https://climate.northwestknowledge.net/MACA/)'
        ws['B5'] = f'Download Date: {download_date}'
        ws['B7'] = f'Location (County): {location_name}'
        ws['B9'] = 'Variable: Jan-Dec (12-month) Standardized Precipitation Evapotranspiration Index (SPEI)'
        ws['B10'] = 'Product: MACAv2-METDATA Downscaled Climate Projections (RCP 4.5)'
        ws['B12'] = '*' * 60
        ws['B13'] = '*' * 60

        # Row 15: Headers
        ws['B15'] = 'Year'
        ws['C15'] = 'CNRM-CM5 (Historical +rcp45)'
        ws['D15'] = 'CanESM2 (Historical +rcp45)'
        ws['E15'] = 'HadGEM2-ES365 (Historical +rcp45)'
        ws['F15'] = 'IPSL-CM5A-MR (Historical +rcp45)'

        # Round values (handle None/NaN gracefully)
        proj_rounded = projections_data.copy()
        for col in ['CNRM-CM5', 'CanESM2', 'HadGEM2-ES365', 'IPSL-CM5A-MR']:
            if col in proj_rounded.columns:
                proj_rounded[col] = pd.to_numeric(proj_rounded[col], errors='coerce').round(2)

        # Row 16+: Data
        for idx, row in proj_rounded.iterrows():
            ws[f'B{16 + idx}'] = int(row['Year'])
            ws[f'C{16 + idx}'] = row.get('CNRM-CM5')
            ws[f'D{16 + idx}'] = row.get('CanESM2')
            ws[f'E{16 + idx}'] = row.get('HadGEM2-ES365')
            ws[f'F{16 + idx}'] = row.get('IPSL-CM5A-MR')

    def _write_dataframe(self, worksheet, df: pd.DataFrame):
        """Write a DataFrame to an Excel worksheet."""
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def weighted_average_dataframes(df1: pd.DataFrame, df2: pd.DataFrame,
                                  weight1: float, weight2: float,
                                  value_cols: list = None) -> pd.DataFrame:
    """
    Calculate weighted average of two DataFrames on 'Year' column.
    """
    if value_cols is None:
        value_cols = [col for col in df1.columns if col != 'Year']

    # Merge on Year
    merged = df1.merge(df2, on='Year', suffixes=('_1', '_2'), how='outer')

    result = pd.DataFrame({'Year': merged['Year']})

    total_weight = weight1 + weight2

    for col in value_cols:
        col1 = f"{col}_1" if f"{col}_1" in merged.columns else col
        col2 = f"{col}_2" if f"{col}_2" in merged.columns else col

        if col1 in merged.columns and col2 in merged.columns:
            # Weighted average where both have data
            result[col] = (merged[col1].fillna(0) * weight1 +
                          merged[col2].fillna(0) * weight2) / total_weight
            # Handle cases where only one has data
            mask1 = merged[col1].notna() & merged[col2].isna()
            mask2 = merged[col1].isna() & merged[col2].notna()
            result.loc[mask1, col] = merged.loc[mask1, col1]
            result.loc[mask2, col] = merged.loc[mask2, col2]
        elif col1 in merged.columns:
            result[col] = merged[col1]
        elif col2 in merged.columns:
            result[col] = merged[col2]

    return result.sort_values('Year').reset_index(drop=True)


def multi_weighted_average(dataframes_weights: list, value_cols: list = None) -> pd.DataFrame:
    """
    Calculate weighted average of multiple DataFrames on 'Year' column.
    dataframes_weights: list of (DataFrame, weight) tuples
    """
    if len(dataframes_weights) == 1:
        return dataframes_weights[0][0]

    if value_cols is None:
        value_cols = [col for col in dataframes_weights[0][0].columns if col != 'Year']

    # Start with first DataFrame
    result = dataframes_weights[0][0].copy()
    total_weight = dataframes_weights[0][1]

    # Merge and weight each subsequent DataFrame
    for df, weight in dataframes_weights[1:]:
        result = result.merge(df, on='Year', how='outer', suffixes=('', '_new'))

        for col in value_cols:
            new_col = f"{col}_new"
            if new_col in result.columns:
                # Weighted combination
                combined = (result[col].fillna(0) * total_weight +
                           result[new_col].fillna(0) * weight) / (total_weight + weight)
                # Handle missing values
                mask_orig = result[col].notna() & result[new_col].isna()
                mask_new = result[col].isna() & result[new_col].notna()
                combined.loc[mask_orig] = result.loc[mask_orig, col]
                combined.loc[mask_new] = result.loc[mask_new, new_col]
                result[col] = combined
                result = result.drop(columns=[new_col])

        total_weight += weight

    return result[['Year'] + value_cols].sort_values('Year').reset_index(drop=True)


def main():
    """Main execution flow."""
    print("=" * 60)
    print("Climate SPEI Data Downloader")
    print("=" * 60)
    print()

    # Collect counties (up to 10)
    county_inputs = []
    max_counties = 10

    for i in range(max_counties):
        if i == 0:
            print(f"Enter county {i+1}:")
        else:
            print()
            add_more = input(f"Add another county? (y/n): ").strip().lower()
            if add_more != 'y':
                break
            print(f"\nEnter county {i+1}:")

        county_name = input("  County name: ").strip()
        state_name = input("  State: ").strip()

        if not county_name or not state_name:
            if i == 0:
                print("Error: At least one county is required.")
                sys.exit(1)
            else:
                print("Skipping empty input.")
                break

        county_inputs.append((county_name, state_name))

    print()
    print("-" * 40)

    # Look up all counties
    county_lookup = CountyLookup()
    counties = []

    for i, (county_name, state_name) in enumerate(county_inputs):
        try:
            county_info = county_lookup.lookup(county_name, state_name)
            area = county_info['geometry'].area
            counties.append((county_info, area))
            print(f"County {i+1}: {county_info['name']} County, {county_info['state']}")
            print(f"  FIPS: {county_info['fips']}, Centroid: {county_info['centroid'][1]:.4f}°N, {abs(county_info['centroid'][0]):.4f}°W")
        except Exception as e:
            print(f"Error looking up {county_name}, {state_name}: {e}")
            if i == 0:
                sys.exit(1)
            print("Skipping this county.")

    if not counties:
        print("Error: No valid counties found.")
        sys.exit(1)

    # Display area-based weighting
    if len(counties) > 1:
        total_area = sum(area for _, area in counties)
        print(f"\nArea-based weighting ({len(counties)} counties):")
        for county_info, area in counties:
            area_sqmi = area * 4639  # Rough conversion at mid-latitudes
            pct = area / total_area * 100
            print(f"  {county_info['name']}: ~{area_sqmi:,.0f} sq mi ({pct:.1f}%)")

    # Track server failures
    server_failures = []

    try:
        wwdt = WestWideDroughtTracker()
        gridmet = GridMETDownloader()
        projections = ClimateProjectionsDownloader()

        if len(counties) == 1:
            # Single county
            county_info = counties[0][0]
            wwdt_data, wwdt_error = wwdt.download(county_info)
            if wwdt_error:
                server_failures.append(("West-Wide Drought Tracker (wrcc.dri.edu)", wwdt_error))

            gridmet_data, gridmet_error = gridmet.download(county_info)
            if gridmet_error:
                server_failures.append(("GridMET (thredds.northwestknowledge.net)", gridmet_error))

            projections_data, failed_models = projections.download(county_info)
            if failed_models:
                server_failures.append(("MACA Climate Projections (tds-proxy.nkn.uidaho.edu)",
                                       f"Failed models: {', '.join(failed_models)}"))

            safe_name = re.sub(r'[^\w\-]', '_', county_info['name'])
            safe_state = county_info['state'][:2].upper()
            output_filename = f"{safe_name}_{safe_state}_SPEI.xlsx"
        else:
            # Multiple counties - weighted average
            print()

            # Download WWDT data for all counties
            wwdt_data_list = []
            for county_info, weight in counties:
                data, error = wwdt.download(county_info)
                if error and not any("West-Wide" in f[0] for f in server_failures):
                    server_failures.append(("West-Wide Drought Tracker (wrcc.dri.edu)", error))
                wwdt_data_list.append((data, weight))
            wwdt_data = multi_weighted_average(wwdt_data_list, ['SPEI'])

            # Download GridMET data for all counties
            gridmet_data_list = []
            for county_info, weight in counties:
                data, error = gridmet.download(county_info)
                if error and not any("GridMET" in f[0] for f in server_failures):
                    server_failures.append(("GridMET (thredds.northwestknowledge.net)", error))
                gridmet_data_list.append((data, weight))
            gridmet_data = multi_weighted_average(gridmet_data_list, ['SPEI'])

            # For projections, use first county
            projections_data, failed_models = projections.download(counties[0][0])
            if failed_models:
                server_failures.append(("MACA Climate Projections (tds-proxy.nkn.uidaho.edu)",
                                       f"Failed models: {', '.join(failed_models)}"))

            # Generate filename from county names
            county_names = [re.sub(r'[^\w\-]', '_', c[0]['name']) for c in counties]
            if len(county_names) <= 3:
                name_part = "_".join(county_names)
            else:
                name_part = f"{county_names[0]}_and_{len(county_names)-1}_others"
            safe_state = counties[0][0]['state'][:2].upper()
            output_filename = f"{name_part}_{safe_state}_SPEI.xlsx"

        output_path = os.path.join("output", output_filename)

        # Build location name for formatted tabs
        if len(counties) == 1:
            location_name = f"{counties[0][0]['name']} County, {counties[0][0]['state']}"
        else:
            county_names = [f"{c[0]['name']}" for c in counties]
            state = counties[0][0]['state']
            if len(county_names) <= 3:
                location_name = f"{', '.join(county_names)} Counties, {state}"
            else:
                location_name = f"{county_names[0]} and {len(county_names)-1} other Counties, {state}"

        # Export to Excel
        exporter = ExcelExporter()
        exporter.export(wwdt_data, gridmet_data, projections_data, output_path, location_name)

        print()
        print("=" * 60)
        print("Download complete!")
        print(f"Output file: {output_path}")

        # Report any server failures
        if server_failures:
            print()
            print("WARNING: Some data servers were unavailable:")
            print("-" * 40)
            for server_name, error_detail in server_failures:
                print(f"  - {server_name}")
                print(f"    Error: {error_detail[:80]}..." if len(error_detail) > 80 else f"    Error: {error_detail}")
            print()
            print("The output file was created but may have missing data.")
            print("Try running again later when servers are back online.")

        print("=" * 60)

    except Exception as e:
        print(f"\nError during processing: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
